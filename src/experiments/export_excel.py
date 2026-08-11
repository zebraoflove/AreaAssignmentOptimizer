from openpyxl import Workbook
from openpyxl.styles import Font


def export_excel(
    conn,
    configuration,
    statistics,
    output_path,
):
    """Экспортирует результаты эксперимента в Excel."""

    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "Summary"

    assignments_sheet = workbook.create_sheet(
        "Assignments"
    )

    write_summary(
        summary_sheet,
        configuration,
        statistics,
    )

    write_assignments(
        assignments_sheet,
        conn,
    )

    workbook.save(output_path)


def write_summary(
    sheet,
    configuration,
    statistics,
):
    """Записывает сводную информацию."""

    sheet["A1"] = "AREA ASSIGNMENT OPTIMIZER"
    sheet["A1"].font = Font(
        bold=True,
        size=16,
    )

    sheet["A3"] = "Algorithm configuration"
    sheet["A3"].font = Font(
        bold=True,
    )

    row = 4

    for key, value in configuration.items():

        sheet.cell(
            row=row,
            column=1,
            value=key,
        )

        sheet.cell(
            row=row,
            column=2,
            value=value,
        )

        row += 1

    row += 1

    sheet.cell(
        row=row,
        column=1,
        value="Assignment statistics",
    ).font = Font(
        bold=True,
    )

    row += 1

    statistics_rows = [
        (
            "Assignments",
            statistics["count"],
        ),
        (
            "Average absolute error",
            statistics["average_absolute_error"],
        ),
        (
            "Average relative error",
            statistics["average_relative_error"],
        ),
    ]

    for key, value in statistics_rows:

        sheet.cell(
            row=row,
            column=1,
            value=key,
        )

        sheet.cell(
            row=row,
            column=2,
            value=value,
        )

        row += 1

    row += 1

    sheet.cell(
        row=row,
        column=1,
        value="Best match",
    ).font = Font(
        bold=True,
    )

    row += 1

    best_match = statistics["best_match"]

    best_rows = [
        ("Subject", best_match["subject_name"]),
        ("Country", best_match["country_name"]),
        ("Absolute error", best_match["absolute_error"]),
        ("Relative error", best_match["relative_error"]),
    ]

    for key, value in best_rows:

        sheet.cell(
            row=row,
            column=1,
            value=key,
        )

        sheet.cell(
            row=row,
            column=2,
            value=value,
        )

        row += 1

    row += 1

    sheet.cell(
        row=row,
        column=1,
        value="Worst match",
    ).font = Font(
        bold=True,
    )

    row += 1

    worst_match = statistics["worst_match"]

    worst_rows = [
        ("Subject", worst_match["subject_name"]),
        ("Country", worst_match["country_name"]),
        ("Absolute error", worst_match["absolute_error"]),
        ("Relative error", worst_match["relative_error"]),
    ]

    for key, value in worst_rows:

        sheet.cell(
            row=row,
            column=1,
            value=key,
        )

        sheet.cell(
            row=row,
            column=2,
            value=value,
        )

        row += 1

    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 30


def write_assignments(
    sheet,
    conn,
):
    """Записывает назначения."""

    headers = [
        "Subject",
        "Subject area (km²)",
        "Country",
        "Country area (km²)",
        "Absolute error (km²)",
        "Relative error (%)",
    ]

    for column, header in enumerate(
        headers,
        start=1,
    ):

        cell = sheet.cell(
            row=1,
            column=column,
            value=header,
        )

        cell.font = Font(
            bold=True,
        )

    rows = conn.execute(
        """
        SELECT
            s.name AS subject_name,
            s.area_km2 AS subject_area,
            c.name AS country_name,
            c.area_km2 AS country_area
        FROM assignments a
        JOIN assignment_items ai
            ON ai.assignment_id = a.id
        JOIN subjects_final s
            ON s.id = a.subject_id
        JOIN countries_final c
            ON c.id = ai.country_id
        ORDER BY
            s.name
        """
    ).fetchall()

    for row_number, row in enumerate(
        rows,
        start=2,
    ):

        subject_area = row["subject_area"]
        country_area = row["country_area"]

        absolute_error = abs(
            subject_area - country_area
        )

        relative_error = (
            absolute_error /
            subject_area
        ) * 100

        values = [
            row["subject_name"],
            subject_area,
            row["country_name"],
            country_area,
            absolute_error,
            relative_error,
        ]

        for column, value in enumerate(
            values,
            start=1,
        ):

            sheet.cell(
                row=row_number,
                column=column,
                value=value,
            )

    widths = {
        "A": 30,
        "B": 20,
        "C": 30,
        "D": 20,
        "E": 22,
        "F": 20,
    }

    for column, width in widths.items():
        sheet.column_dimensions[column].width = width