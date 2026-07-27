from pathlib import Path
import re

from openpyxl import load_workbook

from src.common.config import ROSSTAT_2025_DIR

from src.common.database import connect

DISTRICTS = {
    "CFO": "Центральный",
    "SZFO": "Северо-Западный",
    "YUFO": "Южный",
    "SKFO": "Северо-Кавказский",
    "PFO": "Приволжский",
    "UFO": "Уральский",
    "SFO": "Сибирский",
    "DFO": "Дальневосточный",
}


def detect_district(path: Path) -> str:

    for code, district in DISTRICTS.items():
        if code in path.name:
            return district

    raise ValueError(path.name)


def detect_type(name: str) -> str:

    if "Республика" in name:
        return "республика"

    if name.endswith("край"):
        return "край"

    if name.endswith("область"):
        return "область"

    if "автономный округ" in name:
        return "автономный округ"

    if "автономная область" in name:
        return "автономная область"

    if name.startswith("г. "):
        return "город федерального значения"

    raise ValueError(name)


def parse_area(text: str) -> float:

    text = str(text)

    m = re.search(r"([0-9]+,[0-9]+)\s*тыс", text)

    if not m:
        raise ValueError(f"Не удалось извлечь площадь: {text}")

    return float(m.group(1).replace(",", ".")) * 1000


def read_subjects():

    subjects = []

    for file in sorted(ROSSTAT_2025_DIR.glob("*.xlsx")):

        district = detect_district(file)

        wb = load_workbook(file, data_only=True)

        for ws in wb.worksheets[2:]:

            name = ws["A1"].value
            area_text = ws["A2"].value

            if not name or not area_text:
                continue
            
            name = str(name).strip()
            name = re.sub(r"([а-яё])область$", r"\1 область", name, flags=re.IGNORECASE)
            area = parse_area(str(area_text))

            subjects.append(
                {
                    "name": name,
                    "type": detect_type(name),
                    "federal_district": district,
                    "area_km2": area,
                }
            )

    return subjects


def save_subjects(subjects):

    conn = connect()

    try:

        conn.execute("DELETE FROM subjects")

        conn.executemany(
            """
            INSERT INTO subjects
            (
                name,
                type,
                federal_district,
                area_km2
            )
            VALUES (?, ?, ?, ?)
            """,
            [
                (
                    s["name"],
                    s["type"],
                    s["federal_district"],
                    s["area_km2"],
                )
                for s in subjects
            ],
        )

        conn.commit()

    finally:

        conn.close()


def main():

    subjects = read_subjects()

    save_subjects(subjects)

    print(f"[OK] Imported {len(subjects)} subjects.")


if __name__ == "__main__":
    main()