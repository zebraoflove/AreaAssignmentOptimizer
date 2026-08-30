"""
Подготовка данных субъектов Российской Федерации.
"""

import re

from openpyxl import load_workbook

from src.common.config import ROSSTAT_2025_DIR
from src.common.database import connect
from src.config.subjects import (
    DISTRICTS,
)
from src.pipeline.subject_additions import (
    apply_manual_subject_additions,
)


def detect_district(path):
    """Определяет федеральный округ по имени файла."""

    for code, district in DISTRICTS.items():
        if code in path.name:
            return district

    raise ValueError(path.name)


def detect_type(name):
    """Определяет тип субъекта РФ."""

    if "Республика" in name:
        return "республика"

    if name.endswith("край"):
        return "край"

    if name.endswith("область"):
        return "область"

    if "автономный округ" in name:
        return "автономный округ"

    if "автономная область" in name:
        return "автономная область"

    if name.startswith("г. "):
        return "город федерального значения"

    raise ValueError(name)


def parse_area(text):
    """Извлекает площадь из текста Росстата."""

    text = str(text)

    match = re.search(
        r"([0-9]+(?:,[0-9]+)?)\s*тыс\.",
        text,
        flags=re.IGNORECASE,
    )

    if not match:
        raise ValueError(
            f"Не удалось извлечь площадь: {text}"
        )

    return float(
        match.group(1).replace(",", ".")
    ) * 1000


def read_subjects():
    """Читает субъекты РФ из файлов Росстата."""

    subjects = []

    for file in sorted(
        ROSSTAT_2025_DIR.glob("*.xlsx")
    ):
        district = detect_district(file)

        wb = load_workbook(
            file,
            data_only=True,
        )

        for ws in wb.worksheets[2:]:
            name = ws["A1"].value
            area_text = ws["A2"].value

            if not name or not area_text:
                continue

            name = str(name).strip()

            name = re.sub(
                r"([а-яё])область$",
                r"\1 область",
                name,
                flags=re.IGNORECASE,
            )

            area = parse_area(
                str(area_text)
            )

            subjects.append(
                {
                    "name": name,
                    "type": detect_type(name),
                    "federal_district": district,
                    "area_km2": area,
                }
            )

    return subjects


def save_subjects(conn, subjects):
    """Сохраняет субъекты в subjects_raw."""

    conn.execute(
        "DELETE FROM subjects_raw"
    )

    conn.executemany(
        """
        INSERT INTO subjects_raw (
            name,
            type,
            federal_district,
            area_km2
        )
        VALUES (?, ?, ?, ?)
        """,
        [
            (
                subject["name"],
                subject["type"],
                subject["federal_district"],
                subject["area_km2"],
            )
            for subject in subjects
        ],
    )


def prepare_subjects(conn):
    """Подготавливает исходные данные субъектов РФ."""

    subjects = read_subjects()
    subjects = apply_manual_subject_additions(
        subjects
    )

    save_subjects(
        conn,
        subjects,
    )

    return len(subjects)


def main():
    conn = connect()

    try:
        count = prepare_subjects(conn)

        conn.commit()

        print(
            f"[OK] Imported {count} subjects."
        )

    finally:
        conn.close()


if __name__ == "__main__":
    main()